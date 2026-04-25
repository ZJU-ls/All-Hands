"""XLSX generator · structured sheet payload → .xlsx bytes via openpyxl.

Spec: docs/specs/2026-04-25-artifact-kinds-roadmap.md § 2.2.

Cell type inference (mirrors openpyxl defaults):
- bool      → BOOLEAN cell
- int/float → NUMBER cell
- str       → INLINE_STRING cell (no formula evaluation)
- None      → empty cell
- everything else → str()

We don't accept openpyxl Formula objects from the agent; if a payload starts
with "=" we still write it as a string so the LLM can't cause sheet recalc
side effects on first open.
"""

from __future__ import annotations

import io
from typing import Any

from allhands.execution.artifact_generators.pdf import ArtifactGenerationError

_MAX_SHEETS = 100
_MAX_ROWS_PER_SHEET = 100_000
_MAX_CELL_CHARS = 32_767  # Excel hard limit per cell


def render_xlsx(*, sheets: list[dict[str, Any]]) -> bytes:
    """Build an .xlsx workbook from a list of sheet specs.

    Each spec is ``{"name": str, "headers": list[str]?, "rows": list[list[Any]]}``.
    Headers are written as the first row in bold; if absent, rows start at A1.
    """
    if not sheets:
        raise ArtifactGenerationError("xlsx requires at least one sheet.")
    if len(sheets) > _MAX_SHEETS:
        raise ArtifactGenerationError(
            f"too many sheets: {len(sheets)} > {_MAX_SHEETS}"
        )

    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import Font  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise ArtifactGenerationError(f"openpyxl unavailable: {exc}") from exc

    wb = Workbook()
    # openpyxl seeds a blank "Sheet" we want to replace with the first
    # caller-supplied sheet to keep the workbook order intact.
    wb.remove(wb.active)

    bold = Font(bold=True)
    used_names: set[str] = set()

    for idx, spec in enumerate(sheets):
        raw_name = str(spec.get("name") or f"Sheet{idx + 1}")
        name = _safe_sheet_name(raw_name, used_names)
        used_names.add(name)
        ws = wb.create_sheet(title=name)

        rows = spec.get("rows") or []
        if not isinstance(rows, list):
            raise ArtifactGenerationError(f"sheet[{idx}].rows must be a list.")
        if len(rows) > _MAX_ROWS_PER_SHEET:
            raise ArtifactGenerationError(
                f"sheet[{idx}] exceeds row cap: {len(rows)} > {_MAX_ROWS_PER_SHEET}"
            )

        cursor = 1
        headers = spec.get("headers")
        if headers:
            if not isinstance(headers, list):
                raise ArtifactGenerationError(f"sheet[{idx}].headers must be a list.")
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=cursor, column=col, value=_coerce_cell(h))
                cell.font = bold
            cursor += 1

        for r_offset, row in enumerate(rows):
            if not isinstance(row, list):
                raise ArtifactGenerationError(
                    f"sheet[{idx}].rows[{r_offset}] must be a list."
                )
            for c, val in enumerate(row, start=1):
                ws.cell(row=cursor + r_offset, column=c, value=_coerce_cell(val))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce_cell(val: Any) -> Any:
    """Map agent JSON value → openpyxl-acceptable cell value.

    Strings that start with ``=`` are escaped with a leading apostrophe-
    equivalent (we prepend a zero-width space) so Excel doesn't evaluate
    them as formulas — the agent shouldn't be allowed to inject formulas
    via the structured-build tool.
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val)
    if len(s) > _MAX_CELL_CHARS:
        s = s[: _MAX_CELL_CHARS - 1] + "…"
    if s.startswith("="):
        # Escape so Excel reads the literal text.
        return "'" + s
    return s


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\, no leading/trailing
    apostrophes, must be unique. We trim + suffix-disambiguate without
    raising on collision so an agent that names two sheets the same gets
    a deterministic ``Sheet2`` suffix instead of an error."""
    bad = set("[]:*?/\\")
    cleaned = "".join("_" if c in bad else c for c in name).strip("'")
    cleaned = cleaned[:31] or "Sheet"
    if cleaned not in used:
        return cleaned
    base = cleaned[:28]
    n = 2
    while f"{base}_{n}" in used:
        n += 1
    return f"{base}_{n}"
